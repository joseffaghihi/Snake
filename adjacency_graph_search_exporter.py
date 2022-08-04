import os.path
from multiprocessing import Pool, cpu_count

from numpy import save
from openpyxl import Workbook, load_workbook

import matrix.adjacency_graph_search as ags
from matrix.nodes_to_snake import node_snakes_to_snakes
from snake.snake import remove_reflections, remove_duplicates
from snake.snake_exporter import snakes_to_xlsx


def length_searches(b_max: int, h_max: int, forest: bool):
    if forest:
        path = "data/excel/forest_lengths_bxh.xlsx"
    else:
        path = "data/excel/snake_lengths_bxh.xlsx"
    if os.path.exists(path):
        workbook = load_workbook(path)
    else:
        workbook = Workbook()
    sheet = workbook.active
    sheet.cell(1, 1, "b\\h")
    for h in range(1, h_max + 1):
        sheet.cell(h + 1, 1, h)
    for b in range(1, b_max + 1):
        sheet.cell(1, b + 1, b)
        for h in range(b, h_max + 1):
            if sheet.cell(h + 1, b + 1).value is None:
                print(f"\nFor {b}x{h}")
                if forest:
                    ags.graph_search((b, h), 1, forest, sheet=sheet)
                else:
                    ags.graph_search((b, h), 1, forest)
                sheet.cell(h + 1, b + 1, ags.longuest_snake_length)
                workbook.save(path)
    workbook.close()


def single_search_exporting_all_maximums(b: int, h: int, forest: bool, max_length=0):
    print(f"{b}x{h}: Searching for {b}x{h}...")
    ags.graph_search((b, h), 100000000, forest, max_length=max_length)

    print(f"{b}x{h}: Converting results to snake objects...")
    snakes = node_snakes_to_snakes(ags.longuest_snakes)
    if forest:
        path = f"data/npy/snake_forest_{b}x{h}.npy"
    else:
        path = f"data/npy/snake_{b}x{h}.npy"
    snakes = remove_duplicates(snakes)
    print(f"{b}x{h}: {len(snakes)} fixed snakes")
    save(path, snakes)

    print(f"{b}x{h}: Removing reflections...")
    snakes = remove_reflections(snakes)
    print(f"{b}x{h}: {len(snakes)} free snakes")
    if forest:
        path = f"data/npy/snake_forest_{b}x{h}_no-reflections.npy"
    else:
        path = f"data/npy/snake_{b}x{h}_no-reflections.npy"
    save(path, snakes)

    print(f"{b}x{h}: Exporting to Excel...")
    snakes_to_xlsx(snakes, f"data/excel/graph_search_forest={forest}_{b}x{h}.xlsx")
    print(f"{b}x{h}: Done")
    return snakes


def count_snakes_and_forest(b_min: int, b_max: int, h_max: int):
    filepath = f"data/excel/snakes_and_forests_totals.xlsx"
    if os.path.exists(filepath):
        workbook = load_workbook(filepath)
        snake_sheet = workbook["Snakes"]
        forest_sheet = workbook["Forests"]
    else:
        workbook = Workbook()
        workbook.remove(workbook.active)
        snake_sheet = workbook.create_sheet("Snakes")
        forest_sheet = workbook.create_sheet("Forests")

    snake_sheet.cell(1, 1, "b\\h")
    forest_sheet.cell(1, 1, "b\\h")
    for h in range(1, h_max + 1):
        snake_sheet.cell(h + 1, 1, h)
        forest_sheet.cell(h + 1, 1, h)
    for b in range(1, b_max + 1):
        snake_sheet.cell(1, b + 1, b)
        forest_sheet.cell(1, b + 1, b)

    for b in range(b_min, b_max + 1):
        for h in range(b, h_max + 1):
            snake_done = snake_sheet.cell(h + 1, b + 1).value is not None
            forest_done = forest_sheet.cell(h + 1, b + 1).value is not None
            if not snake_done or not forest_done:
                print(f"\nFor {b}x{h}")
                if not snake_done:
                    ags.graph_search((b, h), 10000000, forest=False)
                    snake_sheet.cell(h + 1, b + 1, len(ags.longuest_snakes))
                if not forest_done:
                    ags.graph_search((b, h), 10000000, forest=True)
                    forest_sheet.cell(h + 1, b + 1, len(ags.longuest_snakes))
                workbook.save(filepath)
    workbook.close()


def loop_search_and_export_to_npy(b_min, b_max, h_max, forest, processes=cpu_count()):
    pool = Pool(processes)
    for b in range(b_min, b_max + 1):
        for h in range(b, h_max + 1):
            if forest:
                path = f"data/npy/snake_forest_{b}x{h}.npy"
            else:
                path = f"data/npy/snake_{b}x{h}.npy"
            if os.path.exists(path) is False:
                pool.apply_async(single_search_exporting_all_maximums, args=(b, h, forest))
    pool.close()
    pool.join()


if __name__ == '__main__':
    # length_searches(10, 40, forest=False)
    # single_search_exporting_all_maximums(5, 23, forest=False)
    loop_search_and_export_to_npy(4, 4, 40, forest=True, processes=5)
    # count_snakes_and_forest(4, 9, 40)
