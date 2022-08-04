import glob
import os.path

import openpyxl


def get_rates_per_epochs():
    results = {x: 0 for x in range(25, 2001, 25)}
    paths = glob.glob(f'{validation_folder_path}/*.png')
    for path in paths:
        path = path[len(f'{validation_folder_path}/'):-len('.png')].split('_')
        if path[0].startswith('+'):
            epoch = int(path[0].strip('+'))
            results[epoch] += 1
    for x in range(25, 2001, 25):
        results[x] = results[x] * 100 / 1000
    return results


def get_rates_per_epochs_and_dimensions(min_rect_dims, max_rect_dims):
    results = {x: {f'{b}x{h}': 0 for b in range(min_rect_dims[0], max_rect_dims[0] + 1) for h in
                   range(min_rect_dims[1], max_rect_dims[1] + 1)} for x in range(25, 2001, 25)}
    total = {x: {f'{b}x{h}': 0 for b in range(min_rect_dims[0], max_rect_dims[0] + 1) for h in
                 range(min_rect_dims[1], max_rect_dims[1] + 1)} for x in range(25, 2001, 25)}
    paths = glob.glob(f'{validation_folder_path}/*.png')
    for path in paths:
        path = path[len(f'{validation_folder_path}/'):-len('.png')].split('_')
        if path[0].startswith('+'):
            epoch = int(path[0].strip('+'))
            results[epoch][path[2]] += 1
            total[epoch][path[2]] += 1
        else:
            epoch = int(path[0])
            total[epoch][path[2]] += 1
    for x in range(25, 2001, 25):
        for dims in results[x].keys():
            results[x][dims] = round(results[x][dims] * 100 / total[x][dims], 2)
    return results


def rates_per_epochs_to_xlsx(rates: dict):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for i, (epoch, rate) in enumerate(rates.items()):
        sheet.cell(i + 1, 1, epoch)
        sheet.cell(i + 1, 2, rate)
    workbook.save(filename=f'{trained_folder_path}/validation_rates_{file_numbers}.xlsx')


def rates_per_epochs_and_dimensions_to_xlsx(rates: dict):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for i, dims in enumerate(rates[25].keys()):
        sheet.cell(1, i + 2, dims)
    for i, (epoch, rate) in enumerate(rates.items()):
        sheet.cell(i + 2, 1, epoch)
        for j, (dims, rate) in enumerate(rate.items()):
            sheet.cell(i + 2, j + 2, rate)
    workbook.save(filename=f'{trained_folder_path}/validation_rates_{file_numbers}.xlsx')


def export_validation_rates_per_epochs():
    rates = get_rates_per_epochs()
    rates_per_epochs_to_xlsx(rates)


def export_validation_rates_per_epochs_and_dimensions():
    rates = get_rates_per_epochs_and_dimensions(min_rect_dims, max_rect_dims)
    rates_per_epochs_and_dimensions_to_xlsx(rates)


if __name__ == '__main__':
    trained_folder_path = f'/home/briseglace/Documents/Serpents/trained_gan'
    for config_number in range(14, 16):
        for gentype in range(4):
            for semtype in range(3):
                min_rect_dims, max_rect_dims = (4, 4), (7, 13)
                file_numbers = f'{config_number}-{gentype}-{semtype}'
                validation_folder_path = f'{trained_folder_path}/{config_number}/gen{gentype}_sem{semtype}/validation'
                if os.path.exists(validation_folder_path):
                    print(f'Processing {file_numbers}...', end=' ')
                    # export_validation_rates_per_epochs()
                    export_validation_rates_per_epochs_and_dimensions()
                    print(f'Done')
