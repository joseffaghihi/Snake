import numpy as np
import openpyxl

from gan.data_images import image_to_snake, snake_to_image


def get_rates(min_rect_dims: tuple, max_rect_dims: tuple, image_dims: tuple):
    rates = \
        {
            f'{b_in}x{h_in}':
                {
                    f'{b_out}x{h_out}': 0
                    for b_out in range(min_rect_dims[0], max_rect_dims[0] + 1)
                    for h_out in range(b_out, max_rect_dims[1] + 1)
                }
            for b_in in range(min_rect_dims[0], max_rect_dims[0] + 1)
            for h_in in range(b_in, max_rect_dims[1] + 1)
        }
    for b_in in range(min_rect_dims[0], max_rect_dims[0] + 1):
        for h_in in range(b_in, max_rect_dims[1] + 1):
            snakes_in = np.load(f"../data/npy/snake_{b_in}x{h_in}.npy", allow_pickle=True)
            for b_out in range(min_rect_dims[0], max_rect_dims[0] + 1):
                for h_out in range(b_out, max_rect_dims[1] + 1):
                    success = 0
                    snakes_out = np.load(f"../data/npy/snake_{b_out}x{h_out}.npy", allow_pickle=True)
                    for snake_in in snakes_in:
                        image = snake_to_image(snake_in, image_dims)
                        converted_bh = image_to_snake(image, (b_out, h_out))
                        converted_hb = image_to_snake(image, (h_out, b_out))
                        for snake_out in snakes_out:
                            if converted_bh == snake_out or converted_hb == snake_out:
                                success += 1
                                break
                    rates[f'{b_in}x{h_in}'][f'{b_out}x{h_out}'] = round(success * 100 / snakes_in.shape[0], 2)
    return rates


def rates_to_xlsx(rates: dict):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for i, dims in enumerate(rates['4x4'].keys()):
        sheet.cell(1, i + 2, dims)
    for i, (dims_in, dict) in enumerate(rates.items()):
        sheet.cell(i + 2, 1, dims_in)
        for j, (dims_out, rate) in enumerate(dict.items()):
            sheet.cell(i + 2, j + 2, rate)
    workbook.save(filename=f'../data/excel/conversions_rates.xlsx')


def export_images_to_snakes_validation(min_rect_dims: tuple, max_rect_dims: tuple, image_dims: tuple):
    rates = get_rates(min_rect_dims, max_rect_dims, image_dims)
    rates_to_xlsx(rates)


if __name__ == '__main__':
    export_images_to_snakes_validation((4, 4), (4, 16), (16, 20))
