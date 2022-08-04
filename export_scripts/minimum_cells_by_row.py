from os import path

from numpy import load
from openpyxl import Workbook

from snake.snake import *


def export_infos(width, height, forest=False):
    workbook = Workbook()
    workbook.remove(workbook.active)

    for b in range(3, width + 1):
        first_column = 1
        sheet = workbook.create_sheet(f'b={b}')
        for h in range(b, height + 1):
            snakes = load_snakes(b, h, forest=forest)
            if len(snakes) > 0:
                print("Processing data...", end=" ")
                results = process_data(snakes)
                first_column = insert_in_sheet(sheet, first_column, results)
                print("Done")
    print("Exporting data...", end=" ")
    if forest:
        workbook.save(filename="../data/excel/forests_bhrn_tmp.xlsx")
    else:
        workbook.save(filename="../data/excel/snakes_bhrn_tmp.xlsx")
    print("Done")


def load_snakes(width, height, forest=False):
    if forest:
        filepath = f"../data/npy/snake_forest_{width}x{height}.npy"
    else:
        filepath = f"../data/npy/snake_{width}x{height}.npy"
    snakes = []
    if path.exists(filepath):
        print(f"Loading {filepath}...", end=" ")
        snakes = load(filepath, allow_pickle=True)
        print("Done")
    return snakes


def search(width, height, processes):
    base_snake = Snake(width=width, height=height)
    snakes = base_snake.find(LONGEST_SNAKES, processes)
    return snakes


def process_data(snakes):
    infos = []
    for snake in snakes:
        infos.append(snake.count_selected_cells_by_row())
    minimums = []
    for r in range(snakes[0].grid.height):
        minimum = infos[0][r]['n']
        for info in infos:
            minimum = min(minimum, info[r]['n'])
        minimums.append({'b': snakes[0].grid.width, 'h': snakes[0].grid.height, 'r': r, 'n': minimum})
    return minimums


def insert_in_sheet(sheet, first_column, min_rows):
    cell_i, cell_j = 1, first_column
    cell = sheet.cell(row=cell_i, column=cell_j)
    cell.value = 'b'
    cell = sheet.cell(row=cell_i, column=cell_j + 1)
    cell.value = 'h'
    cell_i += 1
    cell = sheet.cell(row=cell_i, column=cell_j)
    cell.value = min_rows[0]['b']
    cell = sheet.cell(row=cell_i, column=cell_j + 1)
    cell.value = min_rows[0]['h']
    cell_i += 1
    cell = sheet.cell(row=cell_i, column=cell_j)
    cell.value = 'r'
    cell = sheet.cell(row=cell_i, column=cell_j + 1)
    cell.value = 'n'
    cell_i += 1
    for row in min_rows:
        cell = sheet.cell(row=cell_i, column=cell_j)
        cell.value = row['r']
        cell = sheet.cell(row=cell_i, column=cell_j + 1)
        cell.value = row['n']
        cell_i += 1
    cell_j += 3
    return cell_j


if __name__ == '__main__':
    export_infos(width=8, height=40, forest=False)
