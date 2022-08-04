# PRODUCTION OF AN EXCEL FILE CONTAINING THE NUMBER OF POSSIBLILITIES
# FOR EACH NUMBER OF SNAKES IN FORESTS CONTAINED IN A RECTANGLE.

from os import path

from numpy import save, load
from openpyxl import Workbook

from snake.snake import *


def export_infos(width, height, processes):
    workbook = Workbook()
    sheet = workbook.active
    first_column = 1

    for i in range(1, width + 1):
        for j in range(1, height + 1):
            snakes = get_snakes(i, j, processes)
            print("Processing data...", end=" ")
            data_dict = process_data(snakes)
            first_column = insert_in_sheet(sheet, first_column, i, j, data_dict)
            print("Done")
    print("Exporting data...", end=" ")
    workbook.save(filename="../data/excel/number_of_snakes_forests.xlsx")
    print("Done")


def get_snakes(width, height, processes):
    filepath = f"../data/npy/snake_forest_{width}x{height}.npy"
    if path.exists(filepath):
        print(f"Loading {filepath}...", end=" ")
        snakes = load(filepath, allow_pickle=True)
        print("Done")
    else:
        snakes = search(width, height, processes)
        print("Applying reflections...", end=" ")
        snakes = reflections_mp(snakes=snakes, mode=APPLY, processes=processes)
        print("Done")
        print(f"Saving to {filepath}")
        save(filepath, snakes)
    return snakes


def search(width, height, processes):
    base_snake = Snake(width=width, height=height)
    snakes = base_snake.find(LONGEST_FORESTS, processes)
    return snakes


def process_data(snakes):
    infos = []
    for snake in snakes:
        count = snake.count_snakes()
        infos.append(count)
        infos.sort()
    return {value: infos.count(value) for value in infos}


def insert_in_sheet(sheet, first_column, width, height, data_dict):
    cell_i, cell_j = 1, first_column
    cell = sheet.cell(row=cell_i, column=cell_j)
    cell.value = "Base"
    cell = sheet.cell(row=cell_i, column=cell_j + 1)
    cell.value = "Hauteur"
    cell_i += 1
    cell = sheet.cell(row=cell_i, column=cell_j)
    cell.value = width
    cell = sheet.cell(row=cell_i, column=cell_j + 1)
    cell.value = height
    cell_i += 1
    cell = sheet.cell(row=cell_i, column=cell_j)
    cell.value = "Nombre de serpents"
    cell = sheet.cell(row=cell_i, column=cell_j + 1)
    cell.value = "Nombre de forêts"
    cell_i += 1
    for data in data_dict.items():
        cell = sheet.cell(row=cell_i, column=cell_j)
        cell.value = data[0]
        cell = sheet.cell(row=cell_i, column=cell_j + 1)
        cell.value = data[1]
        cell_i += 1
    cell_j += 3
    return cell_j


if __name__ == '__main__':
    export_infos(width=3, height=7, processes=4)
