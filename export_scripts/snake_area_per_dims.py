from numpy import load
from openpyxl import Workbook

from snake.snake import *
from snake.snake_exporter import snakes_to_xlsx


def export_snakes():
    snakes = get_snakes(3, 6)
    snakes = [snake for snake in snakes if snake.length() == 13]
    snakes_to_xlsx(snakes, '../data/excel/inscribed_snakes_3x6_area13.xlsx')


def export_infos(height):
    workbook = Workbook()
    sheet = workbook.active
    for j in range(width, height + 1):
        snakes = get_snakes(width, j)
        print("Processing data...", end=" ")
        bincount = np.bincount([snake.length() for snake in snakes])
        insert_in_sheet(sheet, j, bincount)
        print("Done")
    print("Exporting data...", end=" ")
    workbook.save(filename=f"../data/excel/inscribed_snakes_area_{width}xh.xlsx")
    print("Done")


def get_snakes(width, height):
    filepath = f"../data/npy/inscribed_snakes_{width}x{height}.npy"
    print(f"Loading {filepath}...", end=" ")
    snakes = load(filepath, allow_pickle=True)
    print("Done")
    return snakes


def insert_in_sheet(sheet, height, bincount):
    cell = sheet.cell(row=1, column=height)
    cell.value = height
    for area, count in enumerate(bincount[1::]):
        cell = sheet.cell(row=area + 2, column=height)
        cell.value = count


if __name__ == '__main__':
    width = 3
    export_infos(15)
    # export_snakes()
