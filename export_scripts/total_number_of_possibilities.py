from os import path

import numpy as np
from openpyxl import Workbook


def main(width, height, forest=False, reflections=False):
    workbook = Workbook()
    workbook.remove(workbook.active)
    sheet = workbook.create_sheet(f'b={width}')
    cell_i, cell_j = 1, 1
    cell = sheet.cell(row=cell_i, column=cell_j)
    cell.value = 'h\\b'
    cell = sheet.cell(row=cell_i, column=cell_j + 1)
    cell.value = width
    for h in range(width, height + 1):
        number = load_number_of_possibilities(width, h, forest=forest, reflections=reflections)
        cell_i += 1
        cell = sheet.cell(row=cell_i, column=cell_j)
        cell.value = h
        cell = sheet.cell(row=cell_i, column=cell_j + 1)
        cell.value = number
    filepath = f"../data/excel/snake_"
    if forest:
        filepath = f"{filepath}forest_"
    filepath = f'{filepath}b{width}'
    if not reflections:
        filepath = f"{filepath}_no-reflections"
    print(f"Saving data to {filepath}.xlsx...", end=" ")
    workbook.save(filename=f"{filepath}.xlsx")
    print("Done")


def load_number_of_possibilities(width, height, forest=False, reflections=False):
    filepath = f"../data/npy/snake_{width}x{height}.npy"
    if forest:
        filepath = f"../data/npy/snake_forest_{width}x{height}.npy"
    if not reflections:
        filepath = filepath[:filepath.find('.npy')] + '_no-reflections.npy'
    number = -1
    if path.exists(filepath):
        print(f"Loading {filepath}...", end=" ")
        number = np.load(filepath, allow_pickle=True).shape[0]
        print("Done")
    return number


if __name__ == '__main__':
    main(width=5, height=40, forest=True, reflections=True)
