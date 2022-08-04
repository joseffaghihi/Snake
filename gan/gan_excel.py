import openpyxl


def export_validation_rates(rates: dict, file_numbers: str):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for i, (epoch, rate) in enumerate(rates.items()):
        sheet.cell(i + 1, 1, epoch)
        sheet.cell(i + 1, 2, rate)
    workbook.save(filename=f'trained/validation_rates_{file_numbers}.xlsx')
