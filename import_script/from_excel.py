from openpyxl import load_workbook


def get_max_length(b, h, forest):
    if forest:
        workbook = load_workbook('data/excel/forest_lengths_bxh.xlsx')
    else:
        workbook = load_workbook('data/excel/snake_lengths_bxh.xlsx')

    return workbook.active.cell(h + 1, b + 1).value
