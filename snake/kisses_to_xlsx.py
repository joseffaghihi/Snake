from numpy import load
from os import path
from openpyxl import Workbook


def export_kisses(width, height):
    workbook = Workbook()
    workbook.remove(workbook.active)
    for b in range(3, width + 1):
        sheet = workbook.create_sheet(f"{b}xh")
        first_column = 1
        for h in range(b, height + 1):
            snakes = get_snakes(b, h)
            if snakes is not None:
                print("Processing data...", end=" ")
                data_dict = process_data(snakes)
                first_column = insert_in_sheet(sheet, first_column, b, h, data_dict)
                print("Done")
    print("Exporting data...", end=" ")
    workbook.save(filename="data/excel/kisses.xlsx")
    print("Done")


def get_snakes(width, height):
    filepath = f"data/npy/snake_{width}x{height}.npy"
    if path.exists(filepath):
        print(f"Loading {filepath}...", end=" ")
        snakes = load(filepath, allow_pickle=True)
        print("Done")
        return snakes


def process_data(snakes):
    infos = []
    for snake in snakes:
        kisses = snake.count_kisses()
        infos.append(kisses)
        infos.sort()
    return {value: infos.count(value) for value in infos}


def insert_in_sheet(sheet, first_column, width, height, data_dict):
    cell_i, cell_j = 1, first_column
    cell = sheet.cell(row=cell_i, column=cell_j + 1)
    cell.value = "Hauteur"
    cell_i += 1
    cell = sheet.cell(row=cell_i, column=cell_j + 1)
    cell.value = height
    cell_i += 1
    cell = sheet.cell(row=cell_i, column=cell_j)
    cell.value = "Kiss"
    cell = sheet.cell(row=cell_i, column=cell_j + 1)
    cell.value = "Serpents"
    cell_i += 1
    for data in data_dict.items():
        cell = sheet.cell(row=cell_i, column=cell_j)
        cell.value = data[0]
        cell = sheet.cell(row=cell_i, column=cell_j + 1)
        cell.value = data[1]
        cell_i += 1
    cell_j += 3
    return cell_j
