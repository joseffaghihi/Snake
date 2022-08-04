from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Border, Side
from snake.snake import Snake


def __choose_borders_style(snake: Snake, i, j) -> Border:
    thin = Side(border_style="thin")
    thick = Side(border_style="thick")
    left, top, right, bottom = thin, thin, thin, thin
    if j == 0:
        left = thick
    if i == 0:
        top = thick
    if i == snake.grid.width - 1:
        bottom = thick
    if j == snake.grid.height - 1:
        right = thick
    return Border(left=left, top=top, right=right, bottom=bottom)


def snakes_to_xlsx(snakes, filepath, sheet=None):
    workbook = None
    if sheet is None:
        workbook = Workbook()
        sheet = workbook.active
    sheet.sheet_format.defaultColWidth = 2.75
    sheet.sheet_format.defaultRowHeight = 12
    green_fill = PatternFill(patternType="solid", fgColor=Color(rgb="00FF00"))
    start_i, start_j = 1, 1
    cell = sheet.cell(row=start_i, column=start_j)
    cell.value = "Longueur"
    cell = sheet.cell(row=start_i, column=start_j + 1)
    cell.value = snakes[0].length()
    start_i += 1
    for snake in snakes:
        for i in range(snake.grid.width):
            for j in range(snake.grid.height):
                cell = sheet.cell(row=start_i + i, column=start_j + j)
                cell.border = __choose_borders_style(snake, i, j)
                if snake.grid.cell_is_selected(i, j):
                    cell.fill = green_fill
        if start_j < 68 - snake.grid.height:
            start_j += snake.grid.height + 1
        else:
            start_i += snake.grid.width + 1
            start_j = 1
    if workbook is not None:
        workbook.save(filename=filepath)


def snake_vs_forest_to_xlsx(snakes, snake_forest, filepath):
    workbook = Workbook()
    workbook.remove(workbook.active)
    sheet = workbook.create_sheet("Serpents maximaux", 0)
    snakes_to_xlsx(snakes, None, sheet)
    sheet = workbook.create_sheet("Ensembles de serpents maximaux", 1)
    snakes_to_xlsx(snake_forest, None, sheet)
    workbook.save(filename=filepath)
